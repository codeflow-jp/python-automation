import os
import datetime

# 現在の日時を取得
now = datetime.datetime.now()
print("現在日時：" + str(now))

# 現在のフォルダのファイル一覧を取得
files = os.listdir(".")
print("ファイル一覧：")
for file in files:
    print("  " + file)

import openpyxl

# 新しいExcelファイルを作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "売上データ"

# ヘッダーを書き込む
ws["A1"] = "名前"
ws["B1"] = "月収"
ws["C1"] = "ステータス"

# データを書き込む
data = [
    ("田中", 350000, "目標達成！"),
    ("佐藤", 120000, "順調です"),
    ("中谷", 50000, "副業軌道中"),
]

for row in data:
    ws.append(row)

# ファイルを保存
wb.save("売上データ.xlsx")
print("Excelファイルを作成しました")